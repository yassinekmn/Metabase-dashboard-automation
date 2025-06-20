# -*- coding: utf-8 -*-
# !pip install playwright reportlab pillow nest_asyncio

# !playwright install chromium

# pip install pandas openpyxl

import os
import asyncio
from playwright.async_api import async_playwright
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import pandas as pd
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
import openpyxl
import glob
# import schedule
import time
from datetime import datetime
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import locale

# Define colors
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_green_fill = PatternFill(start_color="ADFF2F", end_color="ADFF2F", fill_type="solid")  # Yellow-greenish

class MetabaseAgent:
    def __init__(self, metabase_url, username, password):
        self.metabase_url = metabase_url
        self.username = username
        self.password = password
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

        # Define normal and large viewport sizes
        self.normal_viewport = {'width': 1280, 'height': 800}  # Normal viewport size
        self.large_viewport = {'width': 50000, 'height': 50000}  # Larger viewport for table extraction

    async def initialize(self):
        """Initialize Playwright"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=True)
        self.context = await self.browser.new_context()

        # Create a new page after creating the context
        self.page = await self.context.new_page()

        # Set the initial viewport size to the normal size
        await self.page.set_viewport_size(self.normal_viewport)

        # Additional setup (if any)
        await self.context.route('**/*', self._handle_route)

    async def _handle_route(self, route):
        """Handle route interception for capturing API data"""
        await route.continue_()

    async def login(self):
        """Log in to Metabase"""
        try:
            print(f"Navigating to {self.metabase_url}/auth/login")
            await self.page.goto(f"{self.metabase_url}/auth/login")
            print("Filling login form")
            await self.page.fill('input[name="username"]', self.username)
            await self.page.fill('input[name="password"]', self.password)
            print("Submitting login form")
            await self.page.click('button[type="submit"]')
            print("Waiting for navigation element")
            await self.page.wait_for_selector('.Nav', timeout=30000)
            print("Logged in successfully")

            await self.page.screenshot(path="login_success.png")
            print("Saved login success screenshot")
            return True
        except Exception as e:
            print(f"Login failed: {str(e)}")
            await self.page.screenshot(path="login_failed.png")
            print("Saved login failure screenshot")
            return False


    async def wait_until_table_fully_loaded(self, timeout=120, check_interval=1):
        
        start_time = time.time()

        while time.time() - start_time < timeout:
            page_title = await self.page.title()
            # If not loading anymore
            if not re.search(r'\d+/\d+\s+loaded', page_title):
                # But make sure the table is actually visible
                table_exists = await self.page.evaluate('''() => {
                    const table = document.querySelector("table");
                    return table && table.querySelectorAll("tr").length > 0;
                }''')
                if table_exists:
                    print("✅ Table fully loaded and visible.")
                    return True
            print(f"⏳ Still loading: {page_title}")
            await self.page.wait_for_timeout(check_interval * 1000)

        raise TimeoutError("Timed out waiting for table to fully load.")


    async def wait_for_dashboard_to_load(self, timeout=180000):
        """Wait for dashboard to fully load by monitoring the title tag changes"""
        try:
            print("Waiting for dashboard to fully load using title monitoring...")
            start_time = time.time()
            last_reported_progress = None
            final_title_detected = False

            while time.time() - start_time < timeout/1000:
                current_title = await self.page.title()

                # Check if loading pattern has disappeared (final state)
                if not re.search(r'\d+/\d+ loaded', current_title):
                    print("Dashboard fully loaded (loading pattern disappeared)")
                    return True

                # Optional: Report loading progress
                progress_match = re.search(r'(\d+)/(\d+) loaded', current_title)
                if progress_match:
                    current, total = map(int, progress_match.groups())
                    progress = f"{current}/{total}"
                    if progress != last_reported_progress:
                        print(f"Loading progress: {progress}")
                        last_reported_progress = progress

                await self.page.wait_for_timeout(500)  # Check every 500ms

            raise Exception(f"Dashboard did not finish loading within {timeout}ms. Last title: '{await self.page.title()}'")
        except Exception as e:
            print(f"Error waiting for dashboard load: {str(e)}")
            await self.page.screenshot(path="dashboard_load_error.png")
            raise




    async def extract_table_data_to_xlsx(self, card, output_dir, card_id):
        """Extract table data from a table card and save it to an Excel file with clickable first column."""
        try:
            # Before extracting data, set the larger viewport
            await self.page.set_viewport_size(self.large_viewport)
          
            # Try to extract the card title
            card_title = await self.page.evaluate('''(card) => {
                const titleEl = card.querySelector('[data-testid="legend-caption-title"]');
                if (titleEl) return titleEl.textContent.trim();

                const altTitleEl = card.querySelector('.Visualization-title') ||
                                    card.querySelector('.CardVisualization-title') ||
                                    card.querySelector('h3') ||
                                    card.querySelector('.dashcard-title');
                return altTitleEl ? altTitleEl.textContent.trim() : null;
            }''', card)
            
            # Fallback title if not found
            title = card_title if card_title else f"table_{card_id}"
            safe_title = re.sub(r'[\\/*?:"<>|]', "_", title)

            all_rows = []  # List to store all the extracted rows
            headers = []  # To store table headers

            total_pages = 0  # Total number of pages processed
            current_page = 1  # Start with the first page
            
            # Start paginating through the table
            while True:
                print(f"--- Processing page {current_page} ---")
                table_data = await self.page.evaluate('''(card) => {
                    let table = card.querySelector('table');
                    let headers = [];
                    let rows = [];

                    if (!table) {
                        return null;
                    }

                    headers = Array.from(table.querySelectorAll('th')).map(th => th.textContent.trim());
                    const rowsEls = table.querySelectorAll('tr');

                    rows = Array.from(rowsEls).map(tr => {
                        const cells = tr.querySelectorAll('td');
                        // Get both the text content and href attributes for the first cell
                        const rowData = Array.from(cells).map((cell, idx) => {
                            // For the first column, check if there's an anchor tag with href
                            if (idx === 0) {
                                const anchor = cell.querySelector('a');
                                if (anchor && anchor.href) {
                                    return {
                                        text: cell.textContent.trim(),
                                        href: anchor.href
                                    };
                                }
                            }
                            return cell.textContent.trim();
                        });
                        return rowData;
                    }).filter(row => row.length > 0);

                    return { headers, rows };
                }''', card)

                if table_data:
                    if not headers:
                        headers = table_data['headers']

                    all_rows.extend(table_data['rows'])
                    current_page_rows = len(table_data['rows'])
                    print(f"Table on page {current_page} has {current_page_rows} rows")

                current_row_count = len(all_rows)
                print(f"Total rows loaded so far: {current_row_count}")

                # UPDATED PAGINATION DETECTION LOGIC
                has_more_pages = await self.page.evaluate('''() => {
                    // Check if pagination footer exists
                    const tableFooter = document.querySelector('[data-testid="TableFooter"]');
                    if (!tableFooter) return false;
                    
                    // Check if next button exists and is NOT disabled
                    const nextButton = document.querySelector('[aria-label="Page suivante"]');
                    return nextButton && !nextButton.hasAttribute('disabled');
                }''')

                if not has_more_pages:
                    print("No more pages, ending extraction.")
                    total_pages = current_page
                    break

                next_button = await self.page.query_selector('[aria-label="Page suivante"]:not([disabled])')
                if next_button:
                    await next_button.click()
                    print(f"Clicked next page, waiting for table to update...")
                    await self.page.wait_for_timeout(1000)  # Wait for new data to load
                    current_page += 1
                else:
                    print("Next button not found or disabled. Ending extraction.")
                    break

            print(f"Total pages processed: {total_pages}")
            print(f"Total rows extracted: {len(all_rows)}")

            if headers and all_rows:
                # Create a pandas DataFrame for easier data handling
                # But now we need to separate the text and hyperlinks first
                processed_rows = []
                hyperlinks = []

                for row in all_rows:
                    processed_row = []
                    hyperlink = None

                    for idx, cell in enumerate(row):
                        if idx == 0 and isinstance(cell, dict) and 'text' in cell and 'href' in cell:
                            processed_row.append(cell['text'])
                            hyperlink = cell['href']
                        else:
                            processed_row.append(cell)

                    processed_rows.append(processed_row)
                    hyperlinks.append(hyperlink)

                df = pd.DataFrame(processed_rows, columns=headers)
                xlsx_filename = f"{safe_title}.xlsx"
                xlsx_file_path = os.path.join(output_dir, xlsx_filename)

                workbook = openpyxl.Workbook()
                worksheet = workbook.active

                # Add headers
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col_idx, value=header)
                    # Make headers bold
                    worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)

                # Add data rows with hyperlinks in the first column
                for row_idx, (row, hyperlink) in enumerate(zip(processed_rows, hyperlinks), 2):
                    for col_idx, cell in enumerate(row, 1):
                        cell_obj = worksheet.cell(row=row_idx, column=col_idx, value=cell)

                        # Add hyperlink to the first column if available
                        if col_idx == 1 and hyperlink:
                            cell_obj.hyperlink = hyperlink
                            cell_obj.style = "Hyperlink"  # Apply built-in hyperlink style

                # Color the "Nature Intervention" column (index is based on headers)
                if "Nature Intervention" in headers:
                    nature_col_idx = headers.index("Nature Intervention") + 1
                    for row_idx in range(2, len(processed_rows) + 2):
                        nature_value = worksheet.cell(row=row_idx, column=nature_col_idx).value
                        if nature_value == "Curative":
                            worksheet.cell(row=row_idx, column=nature_col_idx).fill = yellow_fill
                        elif nature_value == "Préventive":
                            worksheet.cell(row=row_idx, column=nature_col_idx).fill = green_fill

                # Color the "Date Échéance" column (index is based on headers)
                if "Echéance" in headers:
                    date_echeance_col_idx = headers.index("Echéance") + 1
                    for row_idx in range(2, len(processed_rows) + 2):
                        date_value = worksheet.cell(row=row_idx, column=date_echeance_col_idx).value
                        if isinstance(date_value, str):
                            try:
                                date_value = datetime.strptime(date_value, "%d/%m/%Y")
                            except ValueError:
                                continue  # Skip if the date is in an invalid format

                        if date_value:
                            today = datetime.today()
                            if date_value.date() == today.date():  # It's today
                                worksheet.cell(row=row_idx, column=date_echeance_col_idx).fill = yellow_green_fill
                            elif date_value < today:  # It's in the past
                                worksheet.cell(row=row_idx, column=date_echeance_col_idx).fill = red_fill
                            elif date_value > today:  # It's in the future
                                worksheet.cell(row=row_idx, column=date_echeance_col_idx).fill = green_fill
                
                # Auto-adjust column widths based on content
                for col_idx, header in enumerate(headers, 1):
                    max_length = len(str(header)) + 2  # Add some padding
                    
                    # Check content length in each row for this column
                    for row_idx in range(2, len(processed_rows) + 2):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)) + 2)
                    
                    # Set the column width (max 50 characters to avoid extremely wide columns)
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
                # Save the workbook
                workbook.save(xlsx_file_path)
                print(f"Saved table as Excel with hyperlinks and auto-adjusted columns: {xlsx_file_path}")
                return xlsx_file_path
            else:
                print(f"Table {card_id} has no data.")
                return None

        except Exception as e:
            print(f"Error extracting table data for card {card_id}: {str(e)}")
            traceback.print_exc()  # Print full traceback for better debugging
            return None
        finally:
            # Reset the viewport back to normal size after extraction
            await self.page.set_viewport_size(self.normal_viewport)


    async def extract_dashboard_data(self, dashboard_url, output_dir):
      """Extract data from the dashboard and export it as images and Excel files."""
      os.makedirs(output_dir, exist_ok=True)

      try:
          # Navigate to dashboard
          print(f"Navigating to dashboard URL: {dashboard_url}")
          await self.page.goto(dashboard_url)
          await self.wait_until_table_fully_loaded()
          # Wait for the dashboard to load
          print("Waiting for dashboard grid")
        #   await self.page.wait_for_selector('[data-testid="dashboard-grid"]', timeout=30000)
        #   print("Dashboard grid loaded")

          # Add a longer wait to ensure all data loads
          print("Waiting for data to fully load...")
          await asyncio.sleep(30)  # Add a 5-second delay to allow tables to load

          dashboard_title = await self.page.evaluate('''() => {
              const titleElem = document.querySelector('.Dashboard-header .Entity-title');
              return titleElem ? titleElem.textContent : 'Dashboard';
          }''')
          print(f"Dashboard title: {dashboard_title}")

          dash_cards = await self.page.query_selector_all('.DashCard')
          print(f"Found {len(dash_cards)} dashboard cards")

          card_paths = []
          is_table_card = []

          for idx, card in enumerate(dash_cards):
              card_id = idx + 1
              print(f"Processing card {card_id}/{len(dash_cards)}")

              # Wait for any loading indicators to disappear
              try:
                  # Check if this card has a loading spinner
                  has_loading = await self.page.evaluate('''(card) => {
                      return card.querySelector('.Loading-spinner') !== null ||
                            card.querySelector('[data-testid="loading-spinner"]') !== null ||
                            card.textContent.includes('Loading...');
                  }''', card)

                  if has_loading:
                      print(f"Card {card_id} is still loading, waiting...")
                      # Wait up to 30 seconds for loading to complete
                      for _ in range(30):
                          await asyncio.sleep(1)
                          still_loading = await self.page.evaluate('''(card) => {
                              return card.querySelector('.Loading-spinner') !== null ||
                                    card.querySelector('[data-testid="loading-spinner"]') !== null ||
                                    card.textContent.includes('Loading...');
                          }''', card)
                          if not still_loading:
                              print(f"Card {card_id} finished loading")
                              break
              except Exception as e:
                  print(f"Error checking loading state for card {card_id}: {str(e)}")

              # Enhanced table detection - check for table cells even if the table element isn't directly found
              is_table = await self.page.evaluate('''(card) => {
                  // Try direct table detection first
                  if (card.querySelector('table')) return true;

                  // Check for table-like structures
                  const hasTds = card.querySelectorAll('td').length > 0;
                  const hasThs = card.querySelectorAll('th').length > 0;
                  const hasTableRows = card.querySelectorAll('tr').length > 0;

                  // If we have cells or rows, it's probably a table
                  return (hasTds || hasThs || hasTableRows);
              }''', card)

              # Save this information for later use in PDF generation
              is_table_card.append(is_table)

              # Take screenshot of the card regardless of type
              card_img_path = os.path.join(output_dir, f"card_{card_id}.png")
              await card.screenshot(path=card_img_path)
              print(f"Saved card {card_id} screenshot to {card_img_path}")
              card_paths.append(card_img_path)

              # If it's a table, extract to Excel
              if is_table:
                  print(f"Card {card_id} detected as a table, extracting data...")
                  await self.extract_table_data_to_xlsx(card, output_dir, card_id)
              else:
                  print(f"Card {card_id} is not a table")

          return card_paths, is_table_card

      except Exception as e:
          print(f"Error extracting dashboard data: {str(e)}")
          error_path = os.path.join(output_dir, "error_state.png")
          await self.page.screenshot(path=error_path)
          print(f"Saved error state screenshot to {error_path}")
          raise

    async def close(self):
        """Clean up resources"""
        print("Closing Playwright resources")
        if self.context:
            await self.context.close()
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
        print("Resources closed")



def generate_dashboard_pdf(card_paths, is_table_card, output_pdf_path):
    """Generate PDF with a simpler layout as requested, including pagination."""
    # --- First pass: Count pages ---
    # Filter out table cards
    non_table_cards = [(i, path) for i, (path, is_table) in enumerate(zip(card_paths, is_table_card)) if not is_table]
    
    # --- Layout Settings ---
    width, height = landscape(A4)
    margin = 0.5 * inch
    card_spacing = 0.4 * inch
    page_top_margin = 0.7 * inch
    footer_height = 0.4 * inch  # Space for the footer
    
    # Simulate layout to determine total page count
    def count_total_pages():
        if not non_table_cards:
            return 1  # Just one page for "no cards" message
        
        # First page has title + up to 3 cards in first row
        # Each additional page can fit roughly 4 cards (2 rows of 2)
        remaining_cards = max(0, len(non_table_cards) - 4)
        cards_per_page = 4
        additional_pages = (remaining_cards + cards_per_page - 1) // cards_per_page
        return 1 + additional_pages
    
    total_pages = count_total_pages()
    
    # --- Second pass: Generate the actual PDF ---
    c = canvas.Canvas(output_pdf_path, pagesize=landscape(A4))
    page_num = 1
    
    # Function to draw page border
    def draw_page_border():
        c.setStrokeColorRGB(0, 0, 0)  # Black color
        c.setLineWidth(2)  # Bold line
        c.rect(margin/2, margin/2, width - margin, height - margin)
    
    # Function to draw page number with total
    def draw_page_number():
        c.setFont("Helvetica", 10)
        page_text = f"{page_num}/{total_pages}"
        text_width = c.stringWidth(page_text, "Helvetica", 10)
        c.drawString(width/2 - text_width/2, margin/2 + footer_height/2, page_text)
    
    # Function to draw card border
    def draw_card_border(x, y, w, h):
        c.setStrokeColorRGB(0, 0, 0)  # Black color
        c.setLineWidth(1)  # Thinner line for card borders
        c.rect(x, y, w, h)
    
    # Function to start a new page
    def new_page():
        nonlocal page_num
        draw_page_number()  # Draw page number on current page before moving to next
        c.showPage()
        page_num += 1
        draw_page_border()
        return height - page_top_margin
    
    # Begin actual PDF generation
    # Start first page with border
    draw_page_border()
    y_position = height - page_top_margin
    
    if not non_table_cards:
        c.setFont("Helvetica", 14)
        c.drawString(width/2 - 2*inch, height/2, "No visualization cards to display (tables exported to Excel)")
        draw_page_number()  # Add page number
        c.save()
        print(f"PDF generated with no cards (tables only): {output_pdf_path}")
        return
    
    # --- Generate the title ---
    # --- Title and Date ---
    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
    today_date = datetime.now().strftime("%d %B %Y")  # Example: "06 mai 2025"
    title = f"Rapport quotidien de suivi des demandes en cours\n{today_date}"
    
    # Split the title and date for easier centering
    title_main, title_date = title.split("\n")
    
    # --- Title (Main title + Date) ---
    # Set font to bold and slightly larger size for both title and date
    c.setFont("Helvetica-Bold", 18)
    
    # Draw the main title (first line)
    title_width = c.stringWidth(title_main, "Helvetica-Bold", 18)
    c.drawString(width / 2 - title_width / 2, height - 1 * inch, title_main)
    
    # Draw the date (second line) with reduced space
    date_width = c.stringWidth(title_date, "Helvetica-Bold", 18)
    c.drawString(width / 2 - date_width / 2, height - 1.3 * inch, title_date)  # Reduced space
    
    # --- Adjust position for the next content ---
    y_position -= 1.0 * inch  # Adjust this value to control the space between the title and visualizations
    
    # 2. Next 3 Cards on the same row (if there are enough cards)
    if len(card_paths) > 1:
        card_width = (width - 4 * margin) / 3  # Adjusted width for 3 cards
        card_height = 2 * inch
        
        # Draw the first 3 cards on the same row (start at index 1 since the title was removed)
        for i in range(1, min(4, len(card_paths))):  # Only up to 3 cards (start at index 1)
            img_path = card_paths[i]
            if os.path.exists(img_path):
                x_position = margin + (i - 1) * (card_width + card_spacing)  # Space them out evenly
                c.drawImage(img_path, x_position, y_position - card_height,
                            width=card_width, height=card_height)
                draw_card_border(x_position, y_position - card_height, card_width, card_height)
        
        y_position -= (card_height + card_spacing)  # Move down after the row
    
    # After placing the 3 cards
    remaining_cards = non_table_cards[4:]  # Skip the first 4 cards (1 title + 3 in row)
    
    card_width = (width - 3 * margin) / 2  # Adjust width for 2 cards per row
    card_height = 2.5 * inch
    
    for idx, (orig_idx, img_path) in enumerate(remaining_cards):
        if os.path.exists(img_path):
            # Calculate if we have space for at least one row (2 cards)
            if y_position - (card_height + card_spacing) < margin + footer_height:  # Reserve space for footer
                y_position = new_page()
            
            # Position for 2 cards per row
            x = margin if idx % 2 == 0 else margin + card_width + card_spacing
            c.drawImage(img_path, x, y_position - card_height,
                        width=card_width, height=card_height, preserveAspectRatio=True)
            draw_card_border(x, y_position - card_height, card_width, card_height)
            
            # Move to next row after every second card or if it's the last card in an odd position
            if idx % 2 == 1:
                y_position -= (card_height + card_spacing)
            elif idx == len(remaining_cards) - 1:
                # If it's the last card and it's alone in the row, move down
                y_position -= (card_height + card_spacing)
    
    # Add page number to the last page
    draw_page_number()
    
    # Save the final PDF
    c.save()
    print(f"PDF generated with pagination ({total_pages} pages): {output_pdf_path}")


def get_email_content(fournisseur_name):
    today_date = datetime.now().strftime('%d/%m/%Y')

    # Generate dynamic content for subject, body, and filenames
    subject = f"RAPPORT - {fournisseur_name} ({today_date})"
    body = f"Veuillez trouver le rapport quotidien du {today_date}."
    pdf_filename = f"{fournisseur_name}_situation_{today_date}.pdf"

    excel_filename = f"{fournisseur_name}_rapport_{today_date}.xlsx"

    return subject, body, pdf_filename, excel_filename

def send_report_email(pdf_path, xlsx_files, recipients, subject, fournisseur_name, body, smtp_server, smtp_port,
                     sender_email, sender_password, use_tls=True):
    """
    Send an email with PDF and Excel attachments
    """
    try:
        # Generate email content using fournisseur_name
        subject, body, pdf_filename, excel_filename = get_email_content(fournisseur_name)

        # Create message container
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject

        # Attach the body text
        msg.attach(MIMEText(body, 'plain'))

        # Attach the PDF if it exists
        if pdf_path and os.path.exists(pdf_path):
            with open(pdf_path, 'rb') as file:
                pdf_attachment = MIMEApplication(file.read(), _subtype='pdf')
                pdf_attachment.add_header('Content-Disposition', 'attachment',
                                         filename=pdf_filename)  # Use dynamic filename
                msg.attach(pdf_attachment)
                print(f"Attached PDF: {pdf_filename}")

        # Attach Excel files with dynamic filenames
        for xlsx_path in xlsx_files:
            if os.path.exists(xlsx_path):
                # Extract the title from the path or filename
                table_title = os.path.basename(xlsx_path).replace(".xlsx", "")
                excel_filename = f"{fournisseur_name}_{table_title}_{datetime.now().strftime('%d_%m_%Y')}.xlsx"

                with open(xlsx_path, 'rb') as file:
                    excel_attachment = MIMEApplication(file.read(), _subtype='xlsx')
                    excel_attachment.add_header('Content-Disposition', 'attachment', filename=excel_filename)
                    msg.attach(excel_attachment)
                    print(f"Attached Excel: {excel_filename}")


        server = smtplib.SMTP(smtp_server, smtp_port)
        if use_tls:
            server.starttls()
        server.login(sender_email, sender_password)


        server.sendmail(sender_email, recipients, msg.as_string())
        server.quit()

        print(f"Email sent successfully to {', '.join(recipients)}")
        return True

    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        return False
    
async def process_dashboard(agent, dashboard_url, dashboard_name, output_dir, email_config):
    """Process a single dashboard and send its report via email"""
    dashboard_output_dir = os.path.join(output_dir, dashboard_name)
    os.makedirs(dashboard_output_dir, exist_ok=True)

    try:
        print(f"Extracting data for {dashboard_name}")
        card_paths, is_table_card = await agent.extract_dashboard_data(dashboard_url, dashboard_output_dir)

        # Generate PDF from extracted data (excluding tables)
        print(f"Generating PDF for {dashboard_name}...")
        pdf_output_path = os.path.join(dashboard_output_dir, f"{dashboard_name}_report.pdf")
        generate_dashboard_pdf(card_paths, is_table_card, pdf_output_path)
        print(f"Dashboard PDF created at: {pdf_output_path}")

        # Find all Excel files generated for this dashboard
        xlsx_files = glob.glob(os.path.join(dashboard_output_dir, "*.xlsx"))
        print(f"Found {len(xlsx_files)} Excel files for {dashboard_name}")

        # Send email with the generated files
        print(f"Sending email for {dashboard_name}...")
        today_date = datetime.now().strftime('%d/%m/%Y')
        subject = f"RAPPORT - {dashboard_name} ({today_date})"
        body = f"Veuillez trouver le rapport quotidien du {today_date}."

        email_sent = send_report_email(
            pdf_path=pdf_output_path,
            xlsx_files=xlsx_files,
            recipients=email_config['recipients'],
            subject=subject,
            fournisseur_name=dashboard_name,
            body=body,
            smtp_server=email_config['smtp_server'],
            smtp_port=email_config['smtp_port'],
            sender_email=email_config['sender_email'],
            sender_password=email_config['sender_password']
        )

        if email_sent:
            print(f"Email sent successfully for {dashboard_name}")
        else:
            print(f"Failed to send email for {dashboard_name}")

    except Exception as e:
        print(f"Error processing {dashboard_name}: {str(e)}")




# Load environment variables from .env file
load_dotenv()

async def run_all_dashboards():

    METABASE_URL = os.getenv('METABASE_URL')
    USERNAME = os.getenv('METABASE_USERNAME')
    PASSWORD = os.getenv('METABASE_PASSWORD')
    BASE_OUTPUT_DIR = "dashboard_exports"


    EMAIL_BASE_CONFIG = {
        'smtp_server': os.getenv('SMTP_SERVER'),
        'smtp_port': int(os.getenv('SMTP_PORT')),
        'sender_email': os.getenv('SENDER_EMAIL'),
        'sender_password': os.getenv('SENDER_PASSWORD'),
    }

    today_date = datetime.now().strftime('%d/%m/%Y')

    # Define the three dashboards with their specific configurations
    dashboards = [
        {
            'name': 'PROTECH FM',
            'url': f"{METABASE_URL}/dashboard/2-fournisseur-protech-fm",
            'email': {
                **EMAIL_BASE_CONFIG,
                'recipients': os.getenv('RECIPIENTS').split(','),
                'subject': f"Fournisseur - PROTECH FM",
                'body': f'Veuillez trouver le rapport quotidien du {today_date}.'
            }
        },
         {
            'name': 'SOMEM',
            'url': f"{METABASE_URL}/dashboard/10-fournisseur-somem?",
            'email': {
                **EMAIL_BASE_CONFIG,
                'recipients': os.getenv('RECIPIENTS').split(','),
                'subject': f"Fournisseur - SOMEM ",
                'body': f'Veuillez trouver le rapport quotidien du {today_date}.'
            }
         },
        {
            'name': 'PROCLIM',
            'url': f"{METABASE_URL}/dashboard/4-fournisseur-proclim",
            'email': {
                **EMAIL_BASE_CONFIG,
                'recipients': os.getenv('RECIPIENTS').split(','),
                'subject': f"Fournisseur - PROCLIM",
                'body': f'Veuillez trouver le rapport quotidien du {today_date}.'
            }
        },
        {
            'name': 'SONOTRAB',
            'url': f"{METABASE_URL}/dashboard/3-fournisseur-sonotrab",
            'email': {
                **EMAIL_BASE_CONFIG,
                'recipients': os.getenv('RECIPIENTS').split(','),
                'subject': f"Fournisseur - SONOTRAB",
                'body': f'Veuillez trouver le rapport quotidien du {today_date}.'
            }
        }
    ]

   # Create and initialize agent only once
    agent = MetabaseAgent(METABASE_URL, USERNAME, PASSWORD)
    await agent.initialize()
    login_success = await agent.login()

    if not login_success:
        print("Login failed, aborting all dashboards")
        await agent.close()
        return

    # Process all dashboards using the same logged-in agent
    for dashboard in dashboards:
        print(f"\n--- Processing dashboard: {dashboard['name']} ---\n")
        await process_dashboard(
            agent=agent,
            dashboard_url=dashboard['url'],
            dashboard_name=dashboard['name'],
            output_dir=BASE_OUTPUT_DIR,
            email_config=dashboard['email']
        )

    # Close agent after all dashboards processed
    await agent.close()

import nest_asyncio
nest_asyncio.apply()
asyncio.get_event_loop().run_until_complete(run_all_dashboards())
# import asyncio
# from aiocron import crontab

# @crontab("*/5 * * * *")  # Every 5 minutes
# async def scheduled_task():
#     print(f"\n[{datetime.now()}] Running automated dashboard export")
#     await run_all_dashboards()

# if __name__ == "__main__":
#     import nest_asyncio
#     nest_asyncio.apply()
    
#     print("Scheduler started - will run every 5 minutes")
#     print("Press Ctrl+C to stop")
#     asyncio.get_event_loop().run_forever()
# # ===== End of added code =====